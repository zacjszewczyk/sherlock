# src/formatting.py
import logging
from typing import Dict, List, Tuple, Iterable, Optional
import pandas as pd
from openpyxl.styles import Alignment

logger = logging.getLogger(__name__)

def export_simple_excel(df: pd.DataFrame, path, column_widths_pixels: Optional[Dict[str, int]] = None) -> None:
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="ASOM")
        ws = writer.sheets["ASOM"]
        if column_widths_pixels:
            _apply_pixel_widths(ws, list(df.columns), column_widths_pixels)

def export_with_merged_cells(df: pd.DataFrame, path, column_widths_pixels: Optional[Dict[str, int]] = None) -> None:
    """
    Merge hierarchy spans. If some expected columns are missing (e.g., due to filtering),
    gracefully skip those merge levels with a warning.
    """
    if df.empty:
        logger.warning("export_with_merged_cells: empty DataFrame.")
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="ASOM")
        return

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="ASOM")
        ws = writer.sheets["ASOM"]
        headers = list(df.columns)

        def merge_by_group(group_cols: List[str], merge_cols: List[str], level_name: str) -> None:
            present_group_cols = [c for c in group_cols if c in df.columns]
            present_merge_cols = [c for c in merge_cols if c in df.columns]

            missing_g = [c for c in group_cols if c not in df.columns]
            missing_m = [c for c in merge_cols if c not in df.columns]
            if missing_g or missing_m:
                logger.warning(
                    f"Merging: {level_name}: skipping missing columns "
                    f"(group keys missing={missing_g}, merge cols missing={missing_m})."
                )

            # If there are no group cols or no merge cols present, nothing to do.
            if not present_group_cols or not present_merge_cols:
                return

            # Iterate contiguous spans
            start = 0
            prev = tuple(df.iloc[0][present_group_cols]) if len(df) > 0 else None

            def flush(s: int, e: int):
                if e <= s:
                    return
                top = 2 + s  # header row = 1
                bot = 2 + e
                for mc in present_merge_cols:
                    c = headers.index(mc) + 1
                    ws.merge_cells(start_row=top, start_column=c, end_row=bot, end_column=c)

            for i in range(len(df)):
                key = tuple(df.iloc[i][present_group_cols])
                if key != prev:
                    flush(start, i - 1)
                    start = i
                    prev = key
            flush(start, len(df) - 1)

        # Merge levels (expected full hierarchy)
        merge_by_group(["CCIR Index"], ["CCIR Index", "CCIR"], "CCIR-level")
        merge_by_group(["CCIR Index", "Tactic ID", "Tactic Name"], ["Tactic ID", "Tactic Name"], "Tactic-level")
        merge_by_group(
            ["CCIR Index", "Tactic ID", "Tactic Name", "Indicator Index"],
            ["Indicator Index", "Indicator", "Technique ID", "Technique Name"],
            "Indicator-level"
        )
        merge_by_group(
            ["CCIR Index", "Tactic ID", "Tactic Name", "Indicator Index", "Evidence Index"],
            ["Evidence Index", "Evidence Description", "Data Sources", "Data Platforms", "NAI"],
            "Evidence-level"
        )

        if column_widths_pixels:
            _apply_pixel_widths(ws, headers, column_widths_pixels)

        # Enable wrap text for all populated cells (merged sheet only)
        max_row = ws.max_row
        max_col = ws.max_column
        wrap_align = Alignment(wrap_text=True, vertical="top")
        for r in range(1, max_row + 1):         # include header row if you want header wrapping
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                cell.alignment = wrap_align

def _apply_pixel_widths(ws, headers: List[str], column_widths_pixels: Dict[str, int]) -> None:
    PIXELS_PER_CHAR = 7.0
    from openpyxl.utils import get_column_letter
    for col_name, px in column_widths_pixels.items():
        if col_name not in headers:
            continue
        col_letter = get_column_letter(headers.index(col_name) + 1)
        ch = max(6, int(round(px / PIXELS_PER_CHAR)))
        ws.column_dimensions[col_letter].width = ch

def _col_letter(ws, col_index_1based: int) -> str:
    from openpyxl.utils import get_column_letter
    return get_column_letter(col_index_1based)


def _merge_spans(
    ws,
    df: pd.DataFrame,
    group_on: List[str],
    merge_cols: List[str],
    headers: List[str]
) -> None:
    """
    Find contiguous blocks where 'group_on' fields are identical and merge each of the 'merge_cols'
    across the corresponding row span.
    """
    if any(col not in df.columns for col in group_on):
        return

    # We'll iterate contiguous spans by comparing rows with previous row
    start = 0  # df index (0-based)
    prev_key = tuple(df.iloc[0][group_on]) if not df.empty else None

    def flush_span(s: int, e: int):
        # Merge each column in merge_cols across Excel rows 2+s .. 2+e (header = row 1)
        if e <= s:  # single row - no merge needed
            return
        top_row = 2 + s
        bot_row = 2 + e
        for col in merge_cols:
            if col not in headers:
                continue
            c = headers.index(col) + 1
            ws.merge_cells(start_row=top_row, start_column=c, end_row=bot_row, end_column=c)

    for i in range(len(df)):
        cur_key = tuple(df.iloc[i][group_on])
        if cur_key != prev_key:
            # close previous span [start, i-1]
            flush_span(start, i - 1)
            start = i
            prev_key = cur_key
    # flush last span
    flush_span(start, len(df) - 1)
