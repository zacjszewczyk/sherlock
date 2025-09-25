#!/usr/bin/env python
# coding: utf-8

# # Analytic Scheme of Maneuver Generator
# 
# This notebook uses analytic plans to generate an analytic scheme of maneuver.
# 
# ## Background
# 
# As described in TC 3-12.2.4.1, "The Analytic Scheme of Maneuver is the plan to collect and analyze technical data to meet specific information requirements. It identifies what data to analyze, how to analyze it, and why it is being analyzed." The analytic scheme of maneuver, or ASOM, consists of the following components:
# 
# * **Priority information requirement**:
# * **Indicator**:
# * **Evidence**:
# * **Data**:
# * **NAI**:
# * **Analytic**:
# 
# 
# 
# ## Environment Setup
# 
# This section sets up the environment. It installs packages necessary to generate the analytic plans, imports modules, initializes helper functions, and finally defines global variables. This section also mounts Google Drive to the runtime and moves into the project folder.
# 
# ### Install Packages

# In[1]:


get_ipython().system('pip install -U -q "google" 1> /dev/null')


# ### Import Modules

# In[2]:


from google.colab import userdata
from google.colab import drive
import json
import os
import datetime
import pandas as pd
import re
from collections import defaultdict


# ### Initialize Helper Functions
# 
# The first function, `log`, logs a message to the console prepended with the current timestamp in the ISO8601 format.

# In[3]:


def log(message, end="\n", flush = True):
    """
    Logs a message to the console, prepended with the current timestamp
    in ISO 8601 format.

    Args:
    message (str): The string message to log.
    """

    # Access the global flag controlling verbosity
    global verbose

    # Get the current date and time
    current_time = datetime.datetime.now()

    # Format the timestamp in ISO 8601 format
    timestamp = current_time.isoformat()

    # Construct the final log string using an f-string for clean formatting
    log_string = f"[{timestamp}] {message}"

    # Print the log string to the console if logging is turned on (verbose = True)
    if (verbose == True):
        print(log_string, end = end, flush = flush)


# The second function, `build_asom`, accepts a series of MITRE ATT&CK techniques as input and returns a collection of analytic plans that correspond to those techniques.

# In[4]:


import os
import json
import copy
import re
from pathlib import Path
from typing import Dict, List, Set, Any, Tuple

TACTIC_ID_PATTERN = re.compile(r'^(?P<tactic_id>[A-Z0-9]{2,}-?[A-Z0-9]{0,})')
# Technique IDs: MITRE (T#### or T####.###) OR D3-* style identifiers
TECHNIQUE_ID_PATTERN = re.compile(r'^(?P<tech_id>T\d{4}(?:\.\d{3})?|D3-[A-Z]+)')

def _normalize_tactic_key(tactic: str) -> Tuple[str, str]:
    """
    Accepts strings like:
        "TA0001 - Initial Access"
        "D3-D - Detect"
        "TA0001"
    Returns (tactic_id, tactic_name_or_empty).
    """
    if " - " in tactic:
        tid, name = tactic.split(" - ", 1)
        return tid.strip(), name.strip()
    return tactic.strip(), ""

def _normalize_technique_id(tech: str) -> str:
    """
    Extract technique ID from various accepted forms:
      "T1055"
      "T1055.009"
      "T1055.009 - Process Injection"
      "D3-NTA - Network Traffic Analysis"
      "D3-PM"
    Returns the normalized ID or empty string if not found.
    """
    m = TECHNIQUE_ID_PATTERN.match(tech.strip())
    return m.group("tech_id") if m else ""

def _load_json_safely(path: Path) -> Any:
    """
    Loads JSON; strips code fences if present.
    Returns parsed object or raises.
    """
    text = path.read_text(encoding="utf-8").strip()
    if text.startswith("```"):
        # Remove fenced code markers (``` or ```json)
        text = re.sub(r'^```(?:json)?\s*', '', text)
        text = re.sub(r'\s*```$', '', text)
    return json.loads(text)

def _is_new_schema_object(obj: dict) -> bool:
    required = {"information_requirement", "tactic_id", "tactic_name", "indicators"}
    return isinstance(obj, dict) and required.issubset(obj.keys())

def _filter_indicators(ir_obj: dict, allowed_ids: Set[str]) -> dict:
    """
    Return a *new copy* of ir_obj with indicators filtered to those whose technique_id
    is in allowed_ids. If allowed_ids is empty, returns ir_obj unchanged.
    """
    if not allowed_ids:
        return ir_obj  # no filtering requested

    new_obj = copy.deepcopy(ir_obj)
    new_indicators = [
        ind for ind in new_obj.get("indicators", [])
        if _normalize_technique_id(ind.get("technique_id", "")) in allowed_ids
    ]
    new_obj["indicators"] = new_indicators
    return new_obj

def build_asom(
    attack_chain: Dict[str, List[str]],
    directory: str | Path = ".",
    detect_filename: str = "D3-D.json",
    include_detect_first: bool = True,
    filter_indicators: bool = True,
    deduplicate: bool = True
) -> List[dict]:
    """
    Build an ASOM (list of IR objects in the *new* analytic plan schema) filtered
    by the provided attack_chain.

    Parameters
    ----------
    attack_chain : dict
        Mapping of tactic strings -> list of technique strings.
        Tactic strings can be 'TA0001 - Initial Access' or just 'TA0001'.
        Techniques can be:
            'T1078'
            'T1078 - Valid Accounts'
            'T1055.009'
            'T1055.009 - Process Injection'
            'D3-NTA - Network Traffic Analysis'
            'D3-PM'
    directory : str | Path
        Directory containing new-format analytic plan JSON files.
    detect_filename : str
        Special file whose IR objects should be placed first (if include_detect_first = True).
    include_detect_first : bool
        If True, IR objects from detect_filename (if present) are prepended before the rest.
    filter_indicators : bool
        If True, restrict the 'indicators' list in each returned IR object to only those
        techniques explicitly requested for that tactic. If False, include all indicators
        for matching IR objects (as long as tactic matches).
    deduplicate : bool
        If True, remove duplicates (same information_requirement + tactic_id combination).
        Keeps the first occurrence (maintaining order).

    Returns
    -------
    list[dict]
        List of IR objects (each already in the new schema).
    """

    directory = Path(directory)

    # Normalize attack_chain into:
    #   tactic_map: tactic_id -> set(normalized technique IDs)
    tactic_map: Dict[str, Set[str]] = {}
    for tactic_str, technique_list in attack_chain.items():
        tactic_id, _ = _normalize_tactic_key(tactic_str)
        norm_tecs = {_normalize_technique_id(t) for t in technique_list}
        # Remove empties
        norm_tecs = {t for t in norm_tecs if t}
        tactic_map.setdefault(tactic_id, set()).update(norm_tecs)

    results: List[dict] = []

    def process_file(path: Path):
        try:
            data = _load_json_safely(path)
        except Exception as e:  # noqa
            print(f"[WARN] Could not parse {path.name}: {e}")
            return

        if not isinstance(data, list):
            # print(f"[WARN] {path.name} root is not a list. Skipping.")
            return

        for obj in data:
            if not _is_new_schema_object(obj):
                print(f"[WARN] {path.name} does not conform to schema. Skipping.")
                continue  # silently skip non-conforming entries

            tactic_id = obj.get("tactic_id", "").strip()
            if tactic_id not in tactic_map:
                # print(f"[INFO] {tactic_id} not in tactic map. Skipping.")
                continue  # tactic not requested

            # Possibly filter indicators
            if filter_indicators:
                filtered_obj = _filter_indicators(copy.deepcopy(obj), tactic_map[tactic_id])
                # If after filtering there are no indicators (and we requested some), skip.
                if tactic_map[tactic_id] and not filtered_obj.get("indicators"):
                    continue
                results.append(filtered_obj)
            else:
                results.append(copy.deepcopy(obj))

    # 1. Optionally process the special detect_filename first
    if include_detect_first:
        detect_path = directory / detect_filename
        if detect_path.exists():
            process_file(detect_path)
        else:
            print(f"[ERROR] {detect_filename} not found")
            pass

    # 2. Process all other JSON files
    for path in sorted(directory.glob("*.json")):
        if include_detect_first and path.name == detect_filename:
            continue
        process_file(path)

    # 3. Deduplicate (stable) if requested
    if deduplicate:
        seen = set()
        unique: List[dict] = []
        for obj in results:
            key = (obj.get("information_requirement"), obj.get("tactic_id"))
            if key in seen:
                continue
            seen.add(key)
            unique.append(obj)
        results = unique

    return results


# ### Define Global Variables

# In[5]:


# Toggle logging on (verbose = True)/off (verbose = False)
verbose = True
# verbose = False


# In[6]:


# Rate limits: https://ai.google.dev/gemini-api/docs/rate-limits
# Pricing: https://ai.google.dev/gemini-api/docs/pricing
# Usage: https://console.cloud.google.com/apis/api/generativelanguage.googleapis.com/metrics?project=gen-lang-client-0497172401
# Note that this notebook is designed to be run in Google Colab. The line below reads the Gemini API key for AI Studio,
# which is configured in the Secrets tab on the left side of the Colab window.
os.environ["GEMINI_API_KEY"] = userdata.get("GOOGLE_API_KEY")
log("Gemii API key loaded.")


# ### Mount Google Drive

# In[7]:


# Mount Google Drive and move into the Google AI Studio folder
DRIVE_PATH = "/content/drive"
TECHNIQUES_PATH = "/content/drive/MyDrive/Google AI Studio/analytic-plans"

drive.mount(DRIVE_PATH)
log(f"Google Drive mounted to {DRIVE_PATH}")

os.chdir(TECHNIQUES_PATH)
log(f"Changed directory to {TECHNIQUES_PATH}")


# ## Test Analytic Scheme of Maneuver Generation
# 
# This section generates a small analytic scheme of maneuver off of a test attack chain data set.

# In[8]:


# Example attack_chain data structure
attack_chain_data = {
  "D3-D - Detect": [
    "D3-NTA - Network Traffic Analysis",
    "D3-PM - Platform Monitoring"
  ],
  "TA0001 - Initial Access": [
      "T1133 - External Remote Services"
  ],
  "TA0003 - Persistence": [
    "T1078 - Valid Accounts",
    "T1053 - Scheduled Task/Job"
  ],
  "TA0008 - Lateral Movement": [
    "T1021 - Remote Services",
    "T1570 - Lateral Tool Transfer"
  ],
  "TA0011 - Command and Control": [
    "T1071 - Application Layer Protocol"
  ],
  "TA0010 - Exfiltration": [
     "T1041 - Exfiltration Over C2 Channel",
     "T1048 - Exfiltration Over Alternative Protocol",
     "T1567 - Exfiltration Over Web Service"
  ]
}

log("Building ASOM...")
resulting_asom = build_asom(attack_chain_data)
log("Finished.")


# In[ ]:


# print(json.dumps(resulting_asom, indent=2))


# In[9]:


def format_asom(asom_input_list, joiner="; "):
    """
    Format the output from build_asom_new (new analytic plan schema) into a pandas DataFrame.

    Parameters
    ----------
    asom_input_list : list[dict]
        List of IR objects, each with keys:
          - information_requirement (str)
          - tactic_id (str)
          - tactic_name (str)
          - indicators (list of indicator dicts)
              * indicator dict: { "technique_id", "name", "evidence": [ evidence dicts ] }
              * evidence dict: { "description", "data_sources", "data_platforms", "nai", "action" }
          - version
          - date_created
          - last_updated
          - contributors (list)
    joiner : str
        Delimiter to join list fields (data_sources, data_platforms, contributors).

    Returns
    -------
    pandas.DataFrame
        Tabular view with hierarchical-style indices expressed as string columns:
          IR Index -> Indicator Index (IRIndex.TechSubIndex) -> Evidence Index (IRIndex.TechSubIndex.EvidenceSubIndex)
    """
    import pandas as pd

    table_rows = []
    ir_index = 0

    if not isinstance(asom_input_list, list):
        raise TypeError("Expected asom_input_list to be a list produced by build_asom_new.")

    for ir_obj in asom_input_list:
        if not isinstance(ir_obj, dict):
            print(f"[WARN] Skipping non-dict IR object: {ir_obj}")
            continue

        required_ir_keys = {"information_requirement", "tactic_id", "tactic_name", "indicators"}
        if not required_ir_keys.issubset(ir_obj.keys()):
            print(f"[WARN] IR object missing required keys: {ir_obj.keys()}")
            continue

        indicators = ir_obj.get("indicators", [])
        if not indicators:
            # Optionally emit a placeholder row; for now just warn and skip.
            print(f"[INFO] IR '{ir_obj.get('information_requirement')}' has no indicators; skipping.")
            continue

        ir_index += 1
        tactic_id = ir_obj.get("tactic_id", "")
        tactic_name = ir_obj.get("tactic_name", "")
        information_requirement = ir_obj.get("information_requirement", "")
        information_requirement = f"{information_requirement} ({tactic_id} - {tactic_name})"
        version = ir_obj.get("version", "")
        date_created = ir_obj.get("date_created", "")
        last_updated = ir_obj.get("last_updated", "")
        contributors = ir_obj.get("contributors", [])
        contributors_joined = joiner.join(contributors) if isinstance(contributors, list) else str(contributors)

        tech_sub_index = 0
        for indicator in indicators:
            if not isinstance(indicator, dict):
                print(f"[WARN] Skipping non-dict indicator in IR '{information_requirement}': {indicator}")
                continue

            technique_id = indicator.get("technique_id", "")
            technique_name = indicator.get("name", "")
            evidence_list = indicator.get("evidence", [])

            tech_sub_index += 1
            indicator_index_str = f"{ir_index}.{tech_sub_index}"

            if not evidence_list:
                # If no evidence entries, optionally create a placeholder row.
                # (Current behavior: skip but warn.)
                print(f"[INFO] Indicator '{technique_id} - {technique_name}' has no evidence entries.")
                continue

            evidence_sub_index = 0
            for evidence in evidence_list:
                if not isinstance(evidence, dict):
                    print(f"[WARN] Skipping non-dict evidence under technique '{technique_id}': {evidence}")
                    continue

                evidence_sub_index += 1
                evidence_index_str = f"{indicator_index_str}.{evidence_sub_index}"

                description = evidence.get("description", "")
                data_sources = evidence.get("data_sources", [])
                data_platforms = evidence.get("data_platforms", [])
                nai = evidence.get("nai", "")
                action = evidence.get("action", "")

                # Normalize list fields
                if isinstance(data_sources, list):
                    data_sources_joined = joiner.join(data_sources)
                else:
                    data_sources_joined = str(data_sources)

                if isinstance(data_platforms, list):
                    data_platforms_joined = joiner.join(data_platforms)
                else:
                    data_platforms_joined = str(data_platforms)

                row = {
                    "CCIR Index": ir_index,
                    "CCIR": information_requirement,
                    "Tactic ID": tactic_id,
                    "Tactic Name": tactic_name,
                    "Indicator Index": indicator_index_str,
                    "Indicator": f"{technique_id} - {technique_name}",
                    "Technique ID": technique_id,
                    "Technique Name": technique_name,
                    "Evidence Index": evidence_index_str,
                    "Evidence Description": description,
                    "Data Sources": data_sources_joined,
                    "Data Platforms": data_platforms_joined,
                    "NAI": nai,
                    "Action": action
                    # "Version": version,
                    # "Date Created": date_created,
                    # "Last Updated": last_updated,
                    # "Contributors": contributors_joined,
                }
                table_rows.append(row)

    df = pd.DataFrame(table_rows)

    # Ensure all expected columns exist (important if there were zero rows)
    column_order = [
        "CCIR Index",
        "CCIR",
        "Tactic ID",
        "Tactic Name",
        "Indicator Index",
        "Indicator",
        "Technique ID",
        "Technique Name",
        "Evidence Index",
        "Evidence Description",
        "Data Sources",
        "Data Platforms",
        "NAI",
        "Action"
        # "Version",
        # "Date Created",
        # "Last Updated",
        # "Contributors",
    ]
    for col in column_order:
        if col not in df.columns:
            df[col] = pd.NA

    # Optional: sort by hierarchical indices for readability
    if not df.empty:
        df.sort_values(by=["CCIR Index", "Indicator Index", "Evidence Index"], inplace=True, ignore_index=True)

    return df[column_order]


# In[10]:


formatted_df = format_asom(resulting_asom)


# ### Preview Formatted DataFrame

# In[11]:


# To display the full content of cells if they are long
with pd.option_context('display.max_rows', None,
                        'display.max_columns', None,
                        'display.width', 1000,
                        'display.max_colwidth', None):
    display(formatted_df)


# ### Sort and Re-Index DataFrame

# In[14]:


import pandas as pd
import re

def renumber_formatted_df(formatted_df: pd.DataFrame, d3_tactic_id: str = "D3-D") -> pd.DataFrame:
    """
    Reorder and re-index the formatted analytic summary DataFrame produced from the new schema.

    Fixes prior issue where identical CCIR text received different indices by normalizing the CCIR
    string (collapsing internal whitespace, stripping leading/trailing spaces) before assigning indices.

    Steps
    -----
    1. Sort so that rows with Tactic ID == d3_tactic_id appear first, then remaining tactics
       in normal alphanumeric order of Tactic ID.
    2. Rebuild CCIR Index (1, 2, 3, ...) based on first appearance order (post-sort) of each
       normalized CCIR string.
    3. Rebuild Indicator Index inside each CCIR (e.g., 1.1, 1.2, 2.1, ...), preserving the
       first-seen ordering of each unique Indicator within its CCIR.
    4. Rebuild Evidence Index inside each (CCIR, Indicator) (e.g., 1.1.1, 1.1.2, ...), preserving
       first-seen ordering of each unique Evidence Description within that (CCIR, Indicator).
       Duplicate evidence descriptions within the same (CCIR, Indicator) get the same index.
    5. Return the updated DataFrame (sorted) with refreshed indices.

    Parameters
    ----------
    formatted_df : pd.DataFrame
        The original DataFrame (must include columns: 'CCIR', 'Tactic ID', 'Indicator',
        'Evidence Description', 'CCIR Index', 'Indicator Index', 'Evidence Index').
    d3_tactic_id : str
        The tactic ID to force to the top (default "D3-D").

    Returns
    -------
    pd.DataFrame
        Reindexed and sorted DataFrame.
    """

    df = formatted_df.copy()

    # --- 1. Sorting with D3 tactic first ---
    # Create a primary sort key: 0 for D3 tactic, 1 otherwise.
    df["_tactic_primary_key"] = (df["Tactic ID"] != d3_tactic_id).astype(int)

    # Keep original row order to maintain stable ordering within groups when needed
    df["_orig_row"] = range(len(df))

    # Sort: D3 Tactic first, then by Tactic ID (alphanumeric), then CCIR (normalized),
    # then by Indicator, then Evidence Description to cluster logically.
    def _norm_ccir(s: str) -> str:
        # Normalize whitespace in CCIR (collapse internal spaces & strip)
        return re.sub(r"\s+", " ", s).strip()

    df["_CCIR_norm"] = df["CCIR"].map(_norm_ccir)

    df.sort_values(
        by=["_tactic_primary_key", "Tactic ID", "_CCIR_norm", "Indicator", "Evidence Description", "_orig_row"],
        kind="stable",
        inplace=True
    )

    # --- 2. Rebuild CCIR Index based on normalized CCIR text ---
    ccir_index_map = {}
    next_ccir_idx = 0
    new_ccir_indices = []
    for ccir_norm in df["_CCIR_norm"]:
        # print(ccir_norm)
        if ccir_norm not in ccir_index_map:
            next_ccir_idx += 1
            ccir_index_map[ccir_norm] = next_ccir_idx
        new_ccir_indices.append(ccir_index_map[ccir_norm])
    df["CCIR Index"] = new_ccir_indices

    # --- 3. Rebuild Indicator Index within each CCIR (preserve first-seen order) ---
    indicator_index_map = {}  # (ccir_norm, indicator) -> sub index
    ccir_indicator_counters = {}  # ccir_norm -> next sub index
    new_indicator_indices = []

    for ccir_norm, ccir_idx, indicator in zip(df["_CCIR_norm"], df["CCIR Index"], df["Indicator"]):
        key = (ccir_norm, indicator)
        if key not in indicator_index_map:
            ccir_indicator_counters.setdefault(ccir_norm, 0)
            ccir_indicator_counters[ccir_norm] += 1
            indicator_index_map[key] = ccir_indicator_counters[ccir_norm]
        sub_idx = indicator_index_map[key]
        new_indicator_indices.append(f"{ccir_idx}.{sub_idx}")
    df["Indicator Index"] = new_indicator_indices

    # Precompute (ccir_norm, indicator) -> (ccir_idx, indicator_sub_idx) for evidence indexing
    indicator_sub_lookup = {}
    for (ccir_norm, indicator), sub in indicator_index_map.items():
        indicator_sub_lookup[(ccir_norm, indicator)] = sub

    # --- 4. Rebuild Evidence Index inside each (CCIR, Indicator) ---
    evidence_index_map = {}  # (ccir_norm, indicator, evidence_description) -> sub evidence idx
    evidence_counters = {}   # (ccir_norm, indicator) -> next evidence sub idx
    new_evidence_indices = []

    for ccir_norm, indicator, evidence_desc in zip(df["_CCIR_norm"], df["Indicator"], df["Evidence Description"]):
        parent_key = (ccir_norm, indicator)
        evidence_key = (ccir_norm, indicator, evidence_desc)
        if evidence_key not in evidence_index_map:
            evidence_counters.setdefault(parent_key, 0)
            evidence_counters[parent_key] += 1
            evidence_index_map[evidence_key] = evidence_counters[parent_key]
        ccir_idx = ccir_index_map[ccir_norm]
        indicator_sub_idx = indicator_sub_lookup[parent_key]
        evidence_sub_idx = evidence_index_map[evidence_key]
        new_evidence_indices.append(f"{ccir_idx}.{indicator_sub_idx}.{evidence_sub_idx}")
    df["Evidence Index"] = new_evidence_indices

    # --- 5. Final tidy & cleanup ---
    df.sort_values(
        by=["CCIR Index", "Indicator Index", "Evidence Index"],
        kind="stable",
        inplace=True
    )
    df.drop(columns=["_tactic_primary_key", "_orig_row", "_CCIR_norm"], inplace=True)
    df.reset_index(drop=True, inplace=True)

    return df


# ---------- Example usage (uncomment in your notebook) ----------
updated_df = renumber_formatted_df(formatted_df)

# To display the full content of cells if they are long
with pd.option_context('display.max_rows', None,
                        'display.max_columns', None,
                        'display.width', 1000,
                        'display.max_colwidth', None):
    display(updated_df)


# In[13]:


# Export the full ASOM to an Excel file
# updated_df.to_excel("test_asom_full.xlsx")
updated_df.to_excel("asom_2025.08.19.xlsx")


# ### Create Visually Spanned ASOM

# In[ ]:


import pandas as pd
from pathlib import Path

def _compute_hierarchical_spans(df: pd.DataFrame, span_columns):
    """
    Compute hierarchical row spans.

    For each column in span_columns (ordered highest -> lowest), we:
      * Partition the DataFrame into the ranges (row intervals) created by the
        *previous* column's spans.
      * Within each parent range, merge only contiguous identical values.
      * Record span length at the first row of each run; mark subsequent rows with 0.

    Returns
    -------
    spans : dict[col -> list[int]]
        spans[col][i] = 0 means the cell at (i, col) is hidden under a rowspan
        spans[col][i] = k (k>=1) means a rowspan of length k starts at row i.
    """
    spans = {col: [1] * len(df) for col in span_columns}
    # Start with a single parent range covering all rows
    parent_ranges = [(0, len(df))]

    for col in span_columns:
        col_spans = [1] * len(df)
        new_parent_ranges = []
        for (start, end) in parent_ranges:
            i = start
            while i < end:
                val = df.iat[i, df.columns.get_loc(col)]
                j = i + 1
                while j < end and df.iat[j, df.columns.get_loc(col)] == val:
                    j += 1
                run_len = j - i
                if run_len > 1:
                    col_spans[i] = run_len
                    for r in range(i + 1, j):
                        col_spans[r] = 0
                # Even singleton runs become parent ranges for the next column
                new_parent_ranges.append((i, j))
                i = j
        spans[col] = col_spans
        parent_ranges = new_parent_ranges  # constrain next column to these subranges
    return spans


def html_with_rowspan_hier(df: pd.DataFrame, span_columns):
    """
    Hierarchical version: only merges in a lower column if higher-level
    columns are already merged (or identical) across that same block.
    """
    df = df.copy()
    span_columns = [c for c in span_columns if c in df.columns]
    spans = _compute_hierarchical_spans(df, span_columns)

    def esc(x):
        if pd.isna(x):
            return ""
        return (str(x)
                .replace("&", "&amp;")
                .replace("<", "&lt;")
                .replace(">", "&gt;"))

    cols = list(df.columns)
    parts = [
        "<table border='1' cellspacing='0' cellpadding='4' "
        "style='border-collapse:collapse;font-family:Arial, sans-serif;font-size:12px;'>"
    ]
    parts.append("<thead><tr>" + "".join(f"<th>{esc(c)}</th>" for c in cols) + "</tr></thead>")
    parts.append("<tbody>")

    for i in range(len(df)):
        parts.append("<tr>")
        for col in cols:
            if col in span_columns:
                span_val = spans[col][i]
                if span_val == 0:
                    continue  # covered by an earlier row's rowspan
                elif span_val > 1:
                    parts.append(f"<td rowspan='{span_val}' style='vertical-align:top'>{esc(df.at[i, col])}</td>")
                else:
                    parts.append(f"<td>{esc(df.at[i, col])}</td>")
            else:
                parts.append(f"<td>{esc(df.at[i, col])}</td>")
        parts.append("</tr>")
    parts.append("</tbody></table>")
    return "\n".join(parts)


def display_rowspan_hier(updated_df: pd.DataFrame):
    """
    Convenience display using hierarchical rowspan merging.
    """
    from IPython.display import HTML, display
    span_columns = [
        "CCIR Index", "CCIR",
        "Tactic ID", "Tactic Name",
        "Indicator Index", "Indicator",
        "Technique ID", "Technique Name"
    ]
    html = html_with_rowspan_hier(updated_df, span_columns=span_columns)
    display(HTML(html))


def export_with_merged_cells_hier(
    df: pd.DataFrame,
    span_columns,
    path: str | Path,
    sheet_name="Sheet1",
    header_style=None
):
    """
    Excel export with TRUE hierarchical merges (using openpyxl).
    Lower-level merges never cross boundaries defined by higher-level merges.
    """
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, Font

    df = df.copy()
    span_columns = [c for c in span_columns if c in df.columns]
    spans = _compute_hierarchical_spans(df, span_columns)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    cols = list(df.columns)

    # Header
    for c_idx, col in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=c_idx, value=col)
        if header_style:
            if "font" in header_style: cell.font = header_style["font"]
            if "alignment" in header_style: cell.alignment = header_style["alignment"]
        else:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data
    for r_idx, (_, row) in enumerate(df.iterrows(), start=2):
        for c_idx, col in enumerate(cols, start=1):
            ws.cell(row=r_idx, column=c_idx, value="" if pd.isna(row[col]) else row[col])

    # Apply merges hierarchically using precomputed spans
    for col in span_columns:
        c_idx = cols.index(col) + 1
        r = 0
        while r < len(df):
            span_len = spans[col][r]
            excel_row_start = r + 2  # offset for header
            if span_len > 1:
                ws.merge_cells(
                    start_row=excel_row_start,
                    start_column=c_idx,
                    end_row=excel_row_start + span_len - 1,
                    end_column=c_idx
                )
                anchor = ws.cell(row=excel_row_start, column=c_idx)
                anchor.alignment = Alignment(vertical="top", wrap_text=True)
            r += max(span_len, 1)

    # Simple auto-width
    for c_idx, col in enumerate(cols, start=1):
        max_len = max(len(str(ws.cell(row=r, column=c_idx).value or "")) for r in range(1, len(df) + 2))
        ws.column_dimensions[get_column_letter(c_idx)].width = min(max_len + 2, 60)

    path = Path(path)
    wb.save(path)
    return path


# ---------------- Example Usage ----------------
span_cols = [
    "CCIR Index", "CCIR",
    "Tactic ID", "Tactic Name",
    "Indicator Index", "Indicator",
    "Technique ID", "Technique Name"
]

from IPython.display import HTML
HTML(html_with_rowspan_hier(updated_df, span_cols))

display_rowspan_hier(updated_df)

export_with_merged_cells_hier(updated_df, span_cols, "asom_hierarchical.xlsx")


# In[ ]:




